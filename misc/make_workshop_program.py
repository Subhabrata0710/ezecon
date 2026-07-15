import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\subhabrata\Downloads\Topics EZECON 2026- updated BC V5.xlsx')
ws = wb[wb.sheetnames[0]]

def clean_str(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    val_str = val_str.replace("Cliinical", "Clinical").replace("Regestration", "Registration")
    # Clean up spaces
    val_str = re.sub(r'\s+', ' ', val_str)
    return val_str

def format_time(t_str):
    if not t_str:
        return ""
    t_str = t_str.strip().lower()
    # Normalize formats like '9-9.20 am' to '09:00 AM – 09:20 AM'
    # Or '11.40am-1pm' to '11:40 AM – 01:00 PM'
    # Or '12noon-12.20pm' to '12:00 PM – 12:20 PM'
    # Or '5.30 pm onwards' to '05:30 PM onwards'
    
    t_str = t_str.replace("noon", "pm").replace(".", ":")
    
    # Check if 'onwards' is in the string
    if "onwards" in t_str:
        time_part = t_str.split("onwards")[0].strip()
        # format time_part
        formatted = format_single_time(time_part)
        return f"{formatted} onwards"
        
    parts = t_str.split("-")
    if len(parts) != 2:
        return t_str.upper()
        
    start_str, end_str = parts[0].strip(), parts[1].strip()
    
    # Check if pm/am are specified
    start_meridian = ""
    end_meridian = ""
    
    if "pm" in start_str:
        start_meridian = "PM"
    elif "am" in start_str:
        start_meridian = "AM"
        
    if "pm" in end_str:
        end_meridian = "PM"
    elif "am" in end_str:
        end_meridian = "AM"
        
    # If only one has meridian, propagate it
    if not start_meridian and end_meridian:
        start_meridian = end_meridian
    if not end_meridian and start_meridian:
        end_meridian = start_meridian
        
    start_clean = re.sub(r'[^0-9:]', '', start_str)
    end_clean = re.sub(r'[^0-9:]', '', end_str)
    
    def format_part(time_val, meridian):
        if ":" in time_val:
            h, m = time_val.split(":")
        else:
            h = time_val
            m = "00"
        h = int(h)
        m = int(m)
        return f"{h:02d}:{m:02d} {meridian}"
        
    try:
        s_formatted = format_part(start_clean, start_meridian)
        e_formatted = format_part(end_clean, end_meridian)
        return f"{s_formatted} – {e_formatted}"
    except:
        return t_str.upper()

def format_single_time(t_str):
    t_str = t_str.strip().lower()
    meridian = "PM" if "pm" in t_str else "AM"
    clean = re.sub(r'[^0-9:]', '', t_str)
    if ":" in clean:
        h, m = clean.split(":")
    else:
        h = clean
        m = "00"
    h = int(h)
    m = int(m)
    return f"{h:02d}:{m:02d} {meridian}"

# Test formatter
# print(format_time("9-9.20 am"))
# print(format_time("11.40am-1pm"))
# print(format_time("5.30 pm onwards"))

def make_timeline_item(time_str, topic, level, speaker, chair):
    # Determine style modifiers
    style = ""
    title_style = ""
    is_panel = "panel discussion" in level.lower() or "panel" in topic.lower()
    is_break = "lunch" in topic.lower() or "tea" in topic.lower() or "breakfast" in topic.lower() or "inauguration" in topic.lower() or "general body" in topic.lower() or "awards" in topic.lower()
    
    if is_break:
        style = ' style="border-left-color:var(--color-accent);"'
        title_style = ' style="color:#d32f2f;"'
        
    elif is_panel:
        style = ' style="border-left-color:var(--color-gold);"'
        
    time_formatted = format_time(time_str)
    
    html = f'              <div class="timeline-item"{style}>\n'
    html += f'                <div class="timeline-time">{time_formatted}</div>\n'
    
    # Format topic title
    topic_display = topic
    if is_panel and "panel discussion" not in topic.lower():
        topic_display += " <em>(Panel Discussion)</em>"
        
    html += f'                <div class="timeline-title"{title_style}>{topic_display}</div>\n'
    
    # Speaker & Chairpersons
    speaker_info = []
    if speaker:
        label = "Moderator:" if is_panel else "Speaker:"
        speaker_info.append(f'<strong>{label}</strong> {speaker}')
        if level and not is_panel:
            speaker_info[-1] += f' ({level})'
            
    if chair:
        label = "Panelists:" if is_panel else "Chairpersons:"
        speaker_info.append(f'<strong>{label}</strong> {chair}')
        
    if speaker_info:
        html += f'                <div class="timeline-speaker">{"<br>".join(speaker_info)}</div>\n'
        
    html += '              </div>\n'
    return html

# Generate Day 1 Hall A timeline items
day1_a_html = ""
for r in range(8, 29):
    vals = [cell.value for cell in ws[r]]
    topic = clean_str(vals[2])
    time_slot = clean_str(vals[1])
    level = clean_str(vals[3])
    speaker = clean_str(vals[4])
    chair = clean_str(vals[5])
    if topic:
        day1_a_html += make_timeline_item(time_slot, topic, level, speaker, chair)

# Generate Day 1 Hall B timeline items
day1_b_html = ""
for r in range(8, 28):
    vals = [cell.value for cell in ws[r]]
    topic = clean_str(vals[10])
    time_slot = clean_str(vals[9])
    level = clean_str(vals[11])
    speaker = clean_str(vals[12])
    chair = clean_str(vals[13])
    
    if topic == "Hall A" or topic == "Joining Hall A (Keynote & Inauguration)":
        day1_b_html += '              <div class="timeline-item" style="border-left-color:var(--color-accent);">'
        day1_b_html += '                <div class="timeline-time">11:00 AM – 01:00 PM</div>'
        day1_b_html += '                <div class="timeline-title" style="color:#d32f2f;">→ Joining Hall A (Keynote &amp; Inauguration)</div>'
        day1_b_html += '              </div>\n'
    elif topic:
        day1_b_html += make_timeline_item(time_slot, topic, level, speaker, chair)

# Generate Day 2 Hall A timeline items
day2_a_html = ""
for r in range(37, 55):
    vals = [cell.value for cell in ws[r]]
    topic = clean_str(vals[2])
    time_slot = clean_str(vals[1])
    level = clean_str(vals[3])
    speaker = clean_str(vals[4])
    chair = clean_str(vals[5])
    if topic:
        day2_a_html += make_timeline_item(time_slot, topic, level, speaker, chair)

# Generate Day 2 Hall B timeline items
day2_b_html = ""
for r in range(37, 55):
    vals = [cell.value for cell in ws[r]]
    topic = clean_str(vals[10])
    time_slot = clean_str(vals[9])
    level = clean_str(vals[11])
    speaker = clean_str(vals[12])
    chair = clean_str(vals[13])
    
    if topic:
        day2_b_html += make_timeline_item(time_slot, topic, level, speaker, chair)

# Add standard joined awards item for Day 2 Hall B
day2_b_html += '              <div class="timeline-item" style="border-left-color:var(--color-accent);">'
day2_b_html += '                <div class="timeline-time">03:40 PM – 04:00 PM</div>'
day2_b_html += '                <div class="timeline-title" style="color:#d32f2f;">→ Joining Hall A (Awards &amp; Closing Ceremony)</div>'
day2_b_html += '              </div>\n'

print("=== Day 1 Hall A Items Count ===", day1_a_html.count("timeline-item"))
print("=== Day 1 Hall B Items Count ===", day1_b_html.count("timeline-item"))
print("=== Day 2 Hall A Items Count ===", day2_a_html.count("timeline-item"))
print("=== Day 2 Hall B Items Count ===", day2_b_html.count("timeline-item"))

with open('day1_a.html', 'w', encoding='utf-8') as f:
    f.write(day1_a_html)
with open('day1_b.html', 'w', encoding='utf-8') as f:
    f.write(day1_b_html)
with open('day2_a.html', 'w', encoding='utf-8') as f:
    f.write(day2_a_html)
with open('day2_b.html', 'w', encoding='utf-8') as f:
    f.write(day2_b_html)

