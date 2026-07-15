import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\subhabrata\Downloads\Topics EZECON 2026- updated BC V5.xlsx')
ws = wb[wb.sheetnames[0]]

def clean_str(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    val_str = val_str.replace("Cliinical", "Clinical").replace("Regestration", "Registration")
    val_str = re.sub(r'\s+', ' ', val_str)
    return val_str

def parse_time(time_str):
    """
    Parses a time string like '9-9.20 am', '11.40am-1pm', '12noon-12.20pm', '1-2pm'
    Returns (start_mins, end_mins) from 0:00 (midnight).
    """
    time_str = time_str.lower().replace(" ", "").replace(".", ":")
    
    # Handle noon
    time_str = time_str.replace("noon", "pm")
    
    # Check if we have two parts
    parts = time_str.split("-")
    if len(parts) != 2:
        # Default fallback
        if "pm" in time_str:
            return (13*60, 14*60)
        return (9*60, 10*60)
        
    start_str, end_str = parts[0], parts[1]
    
    # Determine am/pm
    # If am/pm is only at the end, propagate to start
    start_pm = False
    end_pm = False
    
    if "pm" in end_str:
        end_pm = True
    if "am" in end_str:
        end_pm = False
        
    if "pm" in start_str:
        start_pm = True
    elif "am" in start_str:
        start_pm = False
    else:
        # If no meridian on start, match end
        start_pm = end_pm
        
    # Clean non-digit characters except colon
    start_clean = re.sub(r'[^0-9:]', '', start_str)
    end_clean = re.sub(r'[^0-9:]', '', end_str)
    
    def to_mins(t_str, is_pm):
        if not t_str:
            return 9*60
        if ":" in t_str:
            h, m = map(int, t_str.split(":"))
        else:
            h = int(t_str)
            m = 0
            
        if h == 12:
            if not is_pm:
                h = 0
        elif is_pm:
            h += 12
            
        return h * 60 + m

    try:
        start_mins = to_mins(start_clean, start_pm)
        end_mins = to_mins(end_clean, end_pm)
        # Fix cross-meridian or logic check
        if end_mins < start_mins:
            end_mins += 12 * 60 # add 12 hours
        return (start_mins, end_mins)
    except:
        return (9*60, 10*60)

# Let's verify parse_time
# print(parse_time('9-9.20 am')) # (540, 560)
# print(parse_time('11.40am-1pm')) # (700, 780)
# print(parse_time('12noon-12.20pm')) # (720, 740)

def load_events(start_row, end_row):
    events_a = []
    events_b = []
    
    for r in range(start_row, end_row + 1):
        vals = [cell.value for cell in ws[r]]
        
        ha_topic = clean_str(vals[2])
        ha_time = clean_str(vals[1])
        if ha_topic and ha_time:
            start_m, end_m = parse_time(ha_time)
            events_a.append({
                'time_str': ha_time,
                'topic': ha_topic,
                'level': clean_str(vals[3]),
                'speaker': clean_str(vals[4]),
                'chair': clean_str(vals[5]),
                'start_m': start_m,
                'end_m': end_m
            })
            
        hb_topic = clean_str(vals[10])
        hb_time = clean_str(vals[9])
        if hb_topic and hb_time:
            start_m, end_m = parse_time(hb_time)
            events_b.append({
                'time_str': hb_time,
                'topic': hb_topic,
                'level': clean_str(vals[11]),
                'speaker': clean_str(vals[12]),
                'chair': clean_str(vals[13]),
                'start_m': start_m,
                'end_m': end_m
            })
            
    return events_a, events_b

def align_events(events_a, events_b):
    # Find all unique transition minutes
    time_points = set()
    for e in events_a + events_b:
        time_points.add(e['start_m'])
        time_points.add(e['end_m'])
        
    sorted_points = sorted(list(time_points))
    intervals = []
    for i in range(len(sorted_points) - 1):
        intervals.append((sorted_points[i], sorted_points[i+1]))
        
    aligned = []
    for start, end in intervals:
        # Find events in A and B that overlap with this interval
        # If an event covers the interval, we include it
        active_a = [e for e in events_a if e['start_m'] <= start and e['end_m'] >= end]
        active_b = [e for e in events_b if e['start_m'] <= start and e['end_m'] >= end]
        
        aligned.append({
            'start_m': start,
            'end_m': end,
            'ha': active_a[0] if active_a else None,
            'hb': active_b[0] if active_b else None
        })
        
    # Let's merge contiguous intervals that have the same events
    merged = []
    for item in aligned:
        if not merged:
            merged.append(item)
            continue
            
        last = merged[-1]
        # If both HA and HB events are identical to the last interval's events, merge them
        if last['ha'] == item['ha'] and last['hb'] == item['hb']:
            last['end_m'] = item['end_m']
        else:
            merged.append(item)
            
    return merged

def mins_to_str(mins):
    h = mins // 60
    m = mins % 60
    meridian = "am" if h < 12 or h == 24 else "pm"
    h_display = h if h <= 12 else h - 12
    if h_display == 0:
        h_display = 12
    return f"{h_display}:{m:02d} {meridian}"

def format_time_slot(start_m, end_m):
    return f"{mins_to_str(start_m)} – {mins_to_str(end_m)}"

# Load Day 1
d1_a, d1_b = load_events(8, 28)
d1_aligned = align_events(d1_a, d1_b)

# Load Day 2
d2_a, d2_b = load_events(37, 54)
d2_aligned = align_events(d2_a, d2_b)

# Helper to render a card
def render_card(hall_name, data):
    if not data:
        return ""
    
    topic = data['topic']
    is_panel = "panel discussion" in data['level'].lower() or "panel discussion" in topic.lower() or "panel" in topic.lower()
    card_class = "hall-card panel-card" if is_panel else "hall-card"
    if "hall a" in hall_name.lower():
        card_class += " hall-a"
    else:
        card_class += " hall-b"
        
    level_badge = f'<span class="level-badge {data["level"].lower().replace(" ", "-")}">{data["level"]}</span>' if data["level"] else ""
    
    speaker_html = ""
    if data['speaker']:
        label = "Moderator:" if is_panel else "Speaker:"
        speaker_html = f'<div class="card-speaker"><strong>{label}</strong> {data["speaker"]}</div>'
        
    chair_html = ""
    if data['chair']:
        label = "Panelists:" if is_panel else "Chairpersons:"
        chair_html = f'<div class="card-chair"><strong>{label}</strong> {data["chair"]}</div>'
        
    return f"""
        <div class="{card_class}">
          <div class="card-header">
            <span class="hall-badge">{hall_name}</span>
            {level_badge}
          </div>
          <h4 class="card-topic">{topic}</h4>
          {speaker_html}
          {chair_html}
        </div>
    """

def render_combined_card(time_str, topic, extra=""):
    return f"""
    <div class="timeline-slot combined-slot">
      <div class="time-header">{time_str}</div>
      <div class="combined-card">
        <h4 class="card-topic">{topic}</h4>
        {f'<div class="card-extra">{extra}</div>' if extra else ""}
      </div>
    </div>
    """

# Generate Day 1 HTML
d1_html = ""
for item in d1_aligned:
    ha = item['ha']
    hb = item['hb']
    time_str = format_time_slot(item['start_m'], item['end_m'])
    
    if ha and ("lunch" in ha['topic'].lower() or "lunch" in ha['topic'].lower()):
        d1_html += render_combined_card(time_str, "🍽️ Lunch Break")
        continue
    if ha and "inauguration" in ha['topic'].lower():
        d1_html += render_combined_card(time_str, "🎉 Inauguration Ceremony")
        continue
    if ha and "annual general body meeting" in ha['topic'].lower():
        d1_html += render_combined_card(time_str, "📋 Annual General Body Meeting — West Bengal Chapter")
        continue
        
    # Check if Hall B merges with Hall A
    if ha and hb and ("hall a" in hb['topic'].lower() or hb['topic'] == ha['topic']):
        extra_info = f"<strong>Speaker:</strong> {ha['speaker']}"
        if ha['chair']:
            extra_info += f"<br><strong>Chairpersons:</strong> {ha['chair']}"
        d1_html += render_combined_card(time_str, ha['topic'], extra_info)
        continue
        
    # If one of the halls has no event, show the other one full width or side-by-side
    if ha and not hb:
        card_a = render_card("Hall A (Advanced)", ha)
        d1_html += f"""
        <div class="timeline-slot">
          <div class="time-header">{time_str}</div>
          <div class="halls-container single-hall-layout">
            {card_a}
          </div>
        </div>
        """
        continue
        
    if hb and not ha:
        card_b = render_card("Hall B (Basics)", hb)
        d1_html += f"""
        <div class="timeline-slot">
          <div class="time-header">{time_str}</div>
          <div class="halls-container single-hall-layout">
            {card_b}
          </div>
        </div>
        """
        continue
        
    # Standard parallel slots
    card_a = render_card("Hall A (Advanced)", ha)
    card_b = render_card("Hall B (Basics)", hb)
    
    d1_html += f"""
    <div class="timeline-slot">
      <div class="time-header">{time_str}</div>
      <div class="halls-container">
        {card_a}
        {card_b}
      </div>
    </div>
    """

# Generate Day 2 HTML
d2_html = ""
for item in d2_aligned:
    ha = item['ha']
    hb = item['hb']
    time_str = format_time_slot(item['start_m'], item['end_m'])
    
    if ha and "lunch" in ha['topic'].lower():
        d2_html += render_combined_card(time_str, "🍽️ Lunch Break")
        continue
    if ha and "awards distribution" in ha['topic'].lower():
        d2_html += render_combined_card(time_str, "🏆 Awards Distribution and Closing Remarks")
        continue
        
    if ha and hb and ("hall a" in hb['topic'].lower() or hb['topic'] == ha['topic']):
        extra_info = f"<strong>Speaker:</strong> {ha['speaker']}"
        if ha['chair']:
            extra_info += f"<br><strong>Chairpersons:</strong> {ha['chair']}"
        d2_html += render_combined_card(time_str, ha['topic'], extra_info)
        continue
        
    if ha and not hb:
        card_a = render_card("Hall A (Advanced)", ha)
        d2_html += f"""
        <div class="timeline-slot">
          <div class="time-header">{time_str}</div>
          <div class="halls-container single-hall-layout">
            {card_a}
          </div>
        </div>
        """
        continue
        
    if hb and not ha:
        card_b = render_card("Hall B (Basics)", hb)
        d2_html += f"""
        <div class="timeline-slot">
          <div class="time-header">{time_str}</div>
          <div class="halls-container single-hall-layout">
            {card_b}
          </div>
        </div>
        """
        continue
        
    card_a = render_card("Hall A (Advanced)", ha)
    card_b = render_card("Hall B (Basics)", hb)
    
    d2_html += f"""
    <div class="timeline-slot">
      <div class="time-header">{time_str}</div>
      <div class="halls-container">
        {card_a}
        {card_b}
      </div>
    </div>
    """

# Read template and write HTML
template = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Program | EZECON 2026</title>
  <meta name="description" content="Scientific Program — EZECON 2026 Emergency Medicine Conference, 8–9 August 2026.">
  <link rel="stylesheet" href="style.css">
  <style>
    .program-tabs {{
      display: flex;
      justify-content: center;
      gap: 15px;
      margin-bottom: 40px;
    }}
    .tab-btn {{
      padding: 12px 30px;
      border: 2px solid var(--color-primary, #1B4F72);
      background: transparent;
      color: var(--color-primary, #1B4F72);
      border-radius: 50px;
      cursor: pointer;
      font-weight: 700;
      font-size: 1rem;
      transition: all 0.3s ease;
      box-shadow: 0 4px 10px rgba(0,0,0,0.05);
    }}
    .tab-btn.active, .tab-btn:hover {{
      background: var(--color-primary, #1B4F72);
      color: #fff;
      transform: translateY(-2px);
      box-shadow: 0 6px 15px rgba(27,79,114,0.3);
    }}
    .program-timeline {{
      display: none;
      max-width: 1100px;
      margin: 0 auto;
    }}
    .program-timeline.active {{
      display: block;
    }}
    
    .day-header {{
      text-align: center;
      margin-bottom: 30px;
    }}
    .day-badge {{
      display: inline-block;
      background: linear-gradient(135deg, var(--color-primary, #1B4F72), #2C3E50);
      color: #fff;
      padding: 8px 24px;
      border-radius: 30px;
      font-weight: 700;
      font-size: 0.9rem;
      letter-spacing: 1px;
      text-transform: uppercase;
      box-shadow: 0 4px 12px rgba(0,0,0,0.1);
    }}
    
    .timeline-slot {{
      display: grid;
      grid-template-columns: 180px 1fr;
      gap: 25px;
      margin-bottom: 25px;
      position: relative;
    }}
    
    .time-header {{
      font-size: 1.05rem;
      font-weight: 800;
      color: var(--color-primary, #1B4F72);
      padding-top: 15px;
      position: relative;
      text-align: right;
    }}
    
    .halls-container {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
    }}
    .halls-container.single-hall-layout {{
      grid-template-columns: 1fr;
    }}
    
    .hall-card {{
      background: #fff;
      border-radius: 12px;
      padding: 20px;
      box-shadow: 0 4px 20px rgba(0,0,0,0.04);
      border-left: 5px solid #2ecc71; /* Basics (Green) */
      transition: all 0.3s ease;
      display: flex;
      flex-direction: column;
      justify-content: space-between;
    }}
    .hall-card:hover {{
      transform: translateY(-3px);
      box-shadow: 0 8px 30px rgba(0,0,0,0.08);
    }}
    
    .hall-card.hall-a {{
      border-left-color: var(--color-primary, #1B4F72); /* Advanced (Blue) */
    }}
    
    .hall-card.panel-card {{
      border-left-color: #e67e22; /* Panel (Orange) */
      background: #fffaf5;
    }}
    
    .card-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 10px;
    }}
    
    .hall-badge {{
      font-size: 0.75rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.5px;
      color: #7f8c8d;
    }}
    
    .level-badge {{
      font-size: 0.7rem;
      font-weight: 700;
      padding: 3px 8px;
      border-radius: 20px;
      text-transform: uppercase;
    }}
    .level-badge.expert {{
      background: #ffebee;
      color: #c62828;
    }}
    .level-badge.intermediate {{
      background: #e1f5fe;
      color: #0288d1;
    }}
    .level-badge.beginner {{
      background: #e8f5e9;
      color: #2e7d32;
    }}
    .level-badge.panel-discussion {{
      background: #ffe0b2;
      color: #e65100;
    }}
    
    .card-topic {{
      font-size: 1.1rem;
      font-weight: 700;
      color: #2c3e50;
      margin: 0 0 12px 0;
      line-height: 1.4;
      flex-grow: 1;
    }}
    
    .card-speaker, .card-chair {{
      font-size: 0.85rem;
      color: #555;
      margin-top: 4px;
    }}
    .card-speaker strong, .card-chair strong {{
      color: #333;
    }}
    
    /* Combined Slot styling */
    .combined-slot {{
      grid-template-columns: 180px 1fr;
    }}
    .combined-card {{
      background: #f8f9fa;
      border-radius: 12px;
      padding: 20px;
      text-align: center;
      border: 1px dashed #cbd5e1;
    }}
    .combined-card .card-topic {{
      margin: 0;
      font-size: 1.15rem;
      color: var(--color-primary, #1B4F72);
    }}
    .card-extra {{
      margin-top: 10px;
      font-size: 0.9rem;
      color: #555;
    }}
    
    @media (max-width: 900px) {{
      .timeline-slot {{
        grid-template-columns: 1fr;
        gap: 10px;
      }}
      .time-header {{
        text-align: left;
        padding-top: 0;
        border-bottom: 2px solid #eee;
        padding-bottom: 5px;
      }}
      .halls-container {{
        grid-template-columns: 1fr;
        gap: 15px;
      }}
      .combined-slot {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div id="marquee-placeholder"></div>
  <div id="nav-placeholder"></div>

  <section class="page-header">
    <div class="container">
      <h1>Scientific Program</h1>
      <div class="divider"></div>
      <p>8–9 August 2026 · AltAir Boutique Hotel, Kolkata</p>
    </div>
  </section>

  <section class="section section-dark">
    <div class="container">
      
      <div class="program-tabs">
        <button class="tab-btn active" data-day="1">Day 1 — 8th August</button>
        <button class="tab-btn" data-day="2">Day 2 — 9th August</button>
      </div>

      <!-- Day 1 -->
      <div class="program-timeline active" data-day="1">
        <div class="day-header">
          <div class="day-badge">Saturday, 8th August 2026</div>
        </div>
        
        {d1_html}
      </div>

      <!-- Day 2 -->
      <div class="program-timeline" data-day="2">
        <div class="day-header">
          <div class="day-badge">Sunday, 9th August 2026</div>
        </div>
        
        {d2_html}
      </div>

    </div>
  </section>

  <section class="section" style="text-align:center;">
    <div class="container animate-on-scroll">
      <p style="margin-top:1rem;"><a href="register.html" class="btn btn-primary">Register Now →</a></p>
    </div>
  </section>

  <div id="footer-placeholder"></div>
  <button class="back-to-top" aria-label="Back to top">↑</button>
  <script src="script.js"></script>
  <script>
    document.addEventListener('DOMContentLoaded', function(){{
      if(typeof initProgramTabs==='function') initProgramTabs();
    }});
  </script>
</body>
</html>"""

with open('program.html', 'w', encoding='utf-8') as f:
    f.write(template)

print("Generated program.html successfully!")
